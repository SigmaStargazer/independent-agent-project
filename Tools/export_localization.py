# -*- coding: utf-8 -*-
"""
v0.23.5 文案导出工具：策划 Excel（UI文案表.xlsx）→ Unity 运行时 json。

流程（策划数据源 → 运行时文件）：
    策划编辑  GameData/Config/UI文案表.xlsx（唯一编辑入口，src/ 外，策划不进工程目录）
        │  python Tools/export_localization.py
        ▼
    运行时读取  Src/IndependentAgentProject/Assets/Resources/UI/strings_*.json

Excel「全部文案」表约定列（1 起）：
    1 序号 | 2 key | 3 模块 | 4 简体中文 | 5 English | 6 备注
    - 表头第 1 行跳过；中文列禁止留空（英文缺失时运行时回退中文）。
    - 文案可含 {0} {1} 占位符；换行用字面 \\n。

导出行为：
    - 生成 strings_ChineseSimplified.json（简体中文）与 strings_English.json（English）。
    - 英文为空 → 省略该 key（运行时回退中文；不能写空串，空串会被 TryGetValue 命中显示空白）。
    - 保留目标 json 中「不在 Excel 里的 key」（如代码内部使用的 resolution_format），
      避免导出把运行时依赖的附加 key 冲掉。
    - key 重复 / 中文为空时打印错误并中止（防止漏配进游戏）。

用法：
    python Tools/export_localization.py
    python Tools/export_localization.py --excel <自定义Excel路径>
"""
from __future__ import annotations

import argparse
import io
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[export_localization] 缺少依赖 openpyxl，请先安装：uv pip install openpyxl 或 pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 仓库根：本文件位于 <repo>/Tools/export_localization.py
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DEFAULT_EXCEL = os.path.join(REPO_ROOT, "GameData", "Config", "UI文案表.xlsx")
OUT_DIR = os.path.join(
    REPO_ROOT,
    "Src", "IndependentAgentProject", "Assets", "Resources", "UI",
)
OUT_FILES = {
    "简体中文": "strings_ChineseSimplified.json",
    "English": "strings_English.json",
}

# 列索引（1 起）：序号 / key / 模块 / 简体中文 / English / 备注
COL_KEY = 2
COL_ZH = 4
COL_EN = 5


def load_rows(excel_path: str):
    """读取 Excel「全部文案」表，返回 [(key, 中文, English)]，跳过表头、空行、key 为空行。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "全部文案" not in wb.sheetnames:
        print(f"[export_localization] Excel 缺少「全部文案」表: {excel_path}", file=sys.stderr)
        sys.exit(1)
    ws = wb["全部文案"]

    rows = []
    seen = set()
    for r in range(2, ws.max_row + 1):  # 第 1 行表头
        key = ws.cell(r, COL_KEY).value
        zh = ws.cell(r, COL_ZH).value
        en = ws.cell(r, COL_EN).value
        if key is None or str(key).strip() == "":
            continue
        key = str(key).strip()
        zh = str(zh).strip() if zh is not None else ""
        en = str(en).strip() if en is not None else ""

        if key in seen:
            print(f"[export_localization] 错误：key 重复 {key!r}（第 {r} 行）", file=sys.stderr)
            sys.exit(1)
        seen.add(key)
        if zh == "":
            print(f"[export_localization] 错误：{key!r} 简体中文为空（第 {r} 行，中文列禁止留空）", file=sys.stderr)
            sys.exit(1)
        rows.append((key, zh, en))
    return rows


def load_extra_keys(existing_path: str):
    """读现有 json，返回 key→值（导出时保留其中「不在 Excel 里」的附加 key，如 resolution_format）。"""
    extras = {}
    if os.path.exists(existing_path):
        try:
            with io.open(existing_path, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                extras = data
        except Exception as e:
            print(f"[export_localization] 警告：读取现有 {existing_path} 失败，忽略额外 key：{e}", file=sys.stderr)
    return extras


def main():
    parser = argparse.ArgumentParser(description="策划 Excel 文案 → Unity 运行时 json")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="文案表 Excel 路径（默认 GameData/Config/UI文案表.xlsx）")
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"[export_localization] Excel 不存在：{args.excel}", file=sys.stderr)
        sys.exit(1)

    rows = load_rows(args.excel)
    by_lang = {"简体中文": {}, "English": {}}
    for key, zh, en in rows:
        by_lang["简体中文"][key] = zh
        # 英文为空 → 省略该 key（运行时回退中文）。注意：不能写空串，空串会被 TryGetValue 命中而显示空白。
        if en:
            by_lang["English"][key] = en

    # 合并保留「不在 Excel 里的额外 key」（各语言分别取各自现有 json）
    for lang, fname in OUT_FILES.items():
        existing = load_extra_keys(os.path.join(OUT_DIR, fname))
        for k, v in existing.items():
            if k not in by_lang[lang]:
                by_lang[lang][k] = v

    os.makedirs(OUT_DIR, exist_ok=True)
    for lang, fname in OUT_FILES.items():
        out = os.path.join(OUT_DIR, fname)
        with io.open(out, "w", encoding="utf-8", newline="\n") as f:
            json.dump(by_lang[lang], f, ensure_ascii=False, indent=2)
            f.write("\n")
        print(f"[export_localization] 已导出 {lang}: {out} ({len(by_lang[lang])} keys)")

    print(f"[export_localization] 完成。Excel: {args.excel}")


if __name__ == "__main__":
    main()
